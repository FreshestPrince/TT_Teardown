from openpyxl import load_workbook
import pandas as pd
def make_dict(workbook):
    rows = 1
    columns = 1
    wb = load_workbook(workbook, read_only=True)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column
    listoflists = []
    while rows <= row_count:
        while columns <= column_count:
            name = sheet.cell(row=rows, column=columns).value
            df = pd.read_excel(workbook)  # can also index sheet by name or fetch all sheets
            mylist = df[name].tolist()
            listoflists.append(mylist)
            columns += 1
        rows += 1
    num1 = 0
    new_list = []
    while num1 < len(listoflists):
        data = listoflists[num1]
        new_list.append({data[0]: data[1:]})
        num1 += 1
    return new_list

# Creates a list of indexes for when a time is free must sle
def duplicates(lst, item):
    lst = list(lst.values())[0]
    return [i for i, x in enumerate(lst) if x == item]

workbook = "Data.xlsx"
data = make_dict(workbook)
data = data[0]
print(data)
